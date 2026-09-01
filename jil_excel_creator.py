#!/usr/bin/env python3
"""
JIL Excel Creator (Python replacement for JobExcelCreator.java)

Simplification over the Java version:
- No Maven / build step / test case needed. Run directly:  python jil_excel_creator.py <input.jil> [output.xlsx]
- Single file, single dependency (openpyxl).
- Richer classification than the Java version (see classify_job / target_platform).

Usage:
    pip install openpyxl
    python jil_excel_creator.py HK_CA_UAT2.txt
    python jil_excel_creator.py HK_CA_UAT2.txt HK_CA_UAT2_output.xlsx
"""

import re
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Regex patterns for JIL fields ---
P_JOB   = re.compile(r"insert_job:\s*(\S+)")
P_TYPE  = re.compile(r"job_type:\s*(\S+)")
P_CMD   = re.compile(r"command:\s*(.+)")
P_BOX   = re.compile(r"box_name:\s*(\S+)")
P_DATE  = re.compile(r"date_conditions:\s*(\S+)")
P_COND  = re.compile(r"condition:\s*(.+)")
P_START = re.compile(r"start_times:\s*(.+)")
P_RUNCAL = re.compile(r"run_calendar:\s*(\S+)")
P_EXCAL  = re.compile(r"exclude_calendar:\s*(\S+)")
P_MACHINE = re.compile(r"machine:\s*(\S+)")
P_ALARM   = re.compile(r"alarm_if_fail:\s*(\S+)")


def classify_job(name: str, command: str) -> str:
    """Functional category based on job name + command (much richer than Java version)."""
    n = (name + " " + command).upper()
    rules = [
        ("Server Start",      ["_START", "STARTWEBLOGIC", "STARTMANAGED", "START_SERVER"]),
        ("Server Stop",       ["_STOP", "STOPWEBLOGIC", "STOPMANAGED", "STOP_SERVER"]),
        ("NGINX/Web Start",   ["NGINX", "APACHE"]),  # refined below by START/STOP
        ("Sleep/Wait",        ["SLEEP"]),
        ("Feed File Check",   ["FEED_CHK", "CHKFEEDFILE", "FEED_CHECK"]),
        ("Data Upload/Load",  ["UPLOAD", "_UPDATE", "_LOAD", "RSFSREF"]),
        ("Report Generation", ["REPORT"]),
        ("NDM File Transfer", ["NDM"]),
        ("SCP File Transfer", ["SCP"]),
        ("Date Rollover",     ["ROLLOVER"]),
        ("Cache Refresh",     ["CACHE"]),
        ("Permission Change", ["PERMCHG", "CHMOD", "_PERM"]),
        ("Purge/Cleanup",     ["PURGE", "CLEANUP", "CLEAN_LOG", "CLEAN LOGS"]),
        ("Consolidation",     ["CONSOLIDATE"]),
        ("DMS Document",      ["DMS", "LODGE"]),
        ("Entitlement",       ["ENTITLEMENT", "ENTITILEMENT", "ENTIT"]),
        ("Batch Processing",  ["BATCH"]),
        ("Test/Interface",    ["TEST", "INTERFACE"]),
    ]
    for category, keywords in rules:
        if any(k in n for k in keywords):
            # refine web start/stop
            if category == "NGINX/Web Start":
                if "STOP" in n:
                    return "NGINX/Web Stop"
                return "NGINX/Web Start"
            return category
    return "Other / Business Job"


def target_platform(category: str, date_condition, start_times, condition, box_name):
    """
    Decide OpenShift target. RICHER than Java (which only used date_conditions 0/1).
    Returns (target, rationale).
    """
    # 1. Lifecycle jobs are eliminated regardless of date_conditions
    if category in ("Server Start", "Server Stop", "NGINX/Web Start", "NGINX/Web Stop"):
        return ("ELIMINATED", "Container lifecycle managed by OpenShift Deployment/probes")
    if category == "Sleep/Wait":
        return ("ELIMINATED", "Replaced by readiness/liveness probes")

    # 2. Jobs with dependencies need an orchestrator (not a plain CronJob)
    if condition:
        return ("Argo Workflow (DAG)", "Has condition dependency; needs orchestrated workflow")

    # 3. Time-scheduled standalone jobs
    if start_times:
        return ("OpenShift CronJob / Argo CronWorkflow", "Fixed schedule via start_times")

    # 4. date_conditions heuristic (what the Java version used) as fallback
    if str(date_condition) == "1":
        return ("OpenShift CronJob (Scheduled)", "date_conditions=1 (scheduled)")
    if str(date_condition) == "0":
        # Inside a box -> triggered by box, becomes a workflow task
        if box_name:
            return ("Argo Workflow task", "date_conditions=0, triggered within BOX")
        return ("On-demand OpenShift Job", "date_conditions=0, event/manually triggered")

    return ("OpenShift Job", "Default containerized job")


def parse_jil(path):
    jobs = {}
    current = None
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            m = P_JOB.search(line)
            if m:
                name = m.group(1)
                t = P_TYPE.search(line)
                current = name
                jobs[name] = {
                    "name": name,
                    "job_type": t.group(1) if t else "",
                    "box_name": "",
                    "command": "",
                    "script": "",
                    "date_conditions": "",
                    "condition": "",
                    "start_times": "",
                    "run_calendar": "",
                    "exclude_calendar": "",
                    "machine": "",
                    "alarm_if_fail": "",
                }
                continue
            if current is None:
                continue
            j = jobs[current]
            if line.startswith("box_name:"):
                mm = P_BOX.search(line); j["box_name"] = mm.group(1) if mm else ""
            elif line.startswith("command:"):
                mm = P_CMD.search(line)
                if mm:
                    cmd = mm.group(1).strip()
                    j["command"] = cmd
                    # first token that looks like a script
                    first = cmd.split()[0] if cmd.split() else cmd
                    j["script"] = os.path.basename(first) if ("/" in first or "\\" in first) else "cmd"
            elif line.startswith("date_conditions:"):
                mm = P_DATE.search(line); j["date_conditions"] = mm.group(1) if mm else ""
            elif line.startswith("condition:"):
                mm = P_COND.search(line); j["condition"] = mm.group(1).strip() if mm else ""
            elif line.startswith("start_times:"):
                mm = P_START.search(line); j["start_times"] = mm.group(1).strip() if mm else ""
            elif line.startswith("run_calendar:"):
                mm = P_RUNCAL.search(line); j["run_calendar"] = mm.group(1) if mm else ""
            elif line.startswith("exclude_calendar:"):
                mm = P_EXCAL.search(line); j["exclude_calendar"] = mm.group(1) if mm else ""
            elif line.startswith("machine:"):
                mm = P_MACHINE.search(line); j["machine"] = mm.group(1) if mm else ""
            elif line.startswith("alarm_if_fail:"):
                mm = P_ALARM.search(line); j["alarm_if_fail"] = mm.group(1) if mm else ""
    return jobs


def write_excel(jobs, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "Job Name", "Job Type", "Box Name", "Category", "Script Name", "Command",
        "date_conditions", "condition (dependency)", "start_times",
        "run_calendar", "exclude_calendar", "machine", "alarm_if_fail",
        "Target Platform", "Migration Rationale",
    ]
    ws.append(headers)

    # style header
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for j in jobs.values():
        category = classify_job(j["name"], j["command"]) if j["job_type"] != "BOX" else "BOX (container)"
        if j["job_type"] == "BOX":
            target, rationale = ("Argo Workflow (DAG)", "BOX maps to workflow orchestration")
        else:
            target, rationale = target_platform(
                category, j["date_conditions"], j["start_times"], j["condition"], j["box_name"]
            )
        ws.append([
            j["name"], j["job_type"], j["box_name"], category, j["script"], j["command"],
            j["date_conditions"], j["condition"], j["start_times"],
            j["run_calendar"], j["exclude_calendar"], j["machine"], j["alarm_if_fail"],
            target, rationale,
        ])

    # auto width (approximate)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    # ---- Summary sheet ----
    s = wb.create_sheet("Summary")
    total = len(jobs)
    box = sum(1 for j in jobs.values() if j["job_type"] == "BOX")
    cmd = sum(1 for j in jobs.values() if j["job_type"] == "CMD")
    cats = {}
    targets = {}
    for j in jobs.values():
        category = classify_job(j["name"], j["command"]) if j["job_type"] != "BOX" else "BOX (container)"
        if j["job_type"] == "BOX":
            t = "Argo Workflow (DAG)"
        else:
            t, _ = target_platform(category, j["date_conditions"], j["start_times"], j["condition"], j["box_name"])
        cats[category] = cats.get(category, 0) + 1
        targets[t] = targets.get(t, 0) + 1

    s.append(["Metric", "Count"])
    s.append(["Total Jobs", total])
    s.append(["BOX Jobs", box])
    s.append(["CMD Jobs", cmd])
    s.append([])
    s.append(["Category", "Count"])
    for k, v in sorted(cats.items(), key=lambda x: -x[1]):
        s.append([k, v])
    s.append([])
    s.append(["Target Platform", "Count"])
    for k, v in sorted(targets.items(), key=lambda x: -x[1]):
        s.append([k, v])
    for cell in s[1]:
        cell.font = Font(bold=True)

    wb.save(out_path)
    print(f"Excel created: {out_path}  ({total} jobs)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python jil_excel_creator.py <input.jil> [output.xlsx]")
        sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(inp)[0] + "_output.xlsx"
    jobs = parse_jil(inp)
    write_excel(jobs, out)


if __name__ == "__main__":
    main()
